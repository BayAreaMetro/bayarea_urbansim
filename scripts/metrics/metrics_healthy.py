# ==============================
# ====== Healthy Metrics =======
# ==============================

import logging, pathlib, shutil
import pandas as pd
import openpyxl
import xlwings
import metrics_utils
import os


M_DRIVE = (
    pathlib.Path("/Volumes/Data/Models") if os.name != "nt" else pathlib.Path("M:/")
)
BOX_DIR = pathlib.Path("E:/Box")


def generate_parks_from_new_development(fbp_id):
    """
    Creates a DataFrame estimating new park acres generated from residential
    development on parcels for a provided model run. This function processes building
    and parcel data to identify residential developments built between 2020 and 2050,
    categorizes parcels into acreage bins, and calculates the estimated park acreage
    generated as a share of parcel size based on predefined assumptions.

    This function reflects part 4 of off-model strategy EN6 - Expand Urban Greening in
    Communities.  Part 4 was introduced between DBP and FBP of PBA50+, so it should only be
    used with a FBP model run.  Parts 1-3 were provided in an Excel work book and are
    incorporated in the subsequent function, expand_urban_greening().
    Args:
        fbp_id (str): The identifier for the FBP model run, used to locate
            the corresponding new buildings summary file.  Should be a FBP run only.
    Returns:
        pandas.DataFrame: A DataFrame containing the following columns:
            - parcel_id (int): The unique identifier for the parcel.
            - county_name (str): The name of the county where the parcel is located.
            - parcel_acres (float): The size of the parcel in acres.
            - parcel_acres_bin (str): The acreage bin category of the parcel.
            - share_devel_for_parks (float): The share of parcel development
              assumed to be allocated for parks.
            - new_park_acres_pt4 (float): The estimated new park acreage
              generated from the parcel.
    """

    BUILDINGS_FILE = os.path.join(
        M_DRIVE,
        "urban_modeling",
        "baus",
        "PBA50Plus",
        "PBA50Plus_FinalBlueprint",
        fbp_id,
        "core_summaries",
        f"{fbp_id}_new_buildings_summary.csv",
    )

    PARCEL_X_CENSUS_FILE = os.path.join(
        M_DRIVE,
        "urban_modeling",
        "baus",
        "BAUS Inputs",
        "basis_inputs",
        "crosswalks",
        "p10_census.csv",
    )

    # Load buildings and crosswalk to counties
    logging.info(f"Loading new building summary from {BUILDINGS_FILE}")
    buildings = pd.read_csv(BUILDINGS_FILE)
    logging.info(f"Loading parcel to census crosswalk from {PARCEL_X_CENSUS_FILE}")
    parcel_x_census = pd.read_csv(PARCEL_X_CENSUS_FILE)

    # Format county fips
    parcel_x_census["county_fips"] = parcel_x_census["countyfp"].apply(
        lambda x: f"06{x:03d}"
    )

    # Map county_fips to county_name
    county_fips_to_name = {
        "06001": "Alameda",
        "06013": "Contra Costa",
        "06041": "Marin",
        "06055": "Napa",
        "06075": "San Francisco",
        "06081": "San Mateo",
        "06085": "Santa Clara",
        "06095": "Solano",
        "06097": "Sonoma",
    }
    parcel_x_census["county_name"] = parcel_x_census["county_fips"].map(
        county_fips_to_name
    )

    # Reduce to required columns
    parcel_x_county = parcel_x_census[["parcel_id", "county_fips", "county_name"]]

    # Merge buildings and parcel_x_county
    logging.info(f"Merging county names to buildings")
    buildings_county = buildings.merge(parcel_x_county, on="parcel_id", how="left")

    # Keep only residential parcels with development between 2020 and 2050 and remove duplicate parcels
    residential_types = ["HS", "HT", "HM", "MR"]
    initial_year = 2020
    logging.info(
        f"Reducing to residential buildings with type {residential_types} built {initial_year} to 2050"
    )
    res_buildings_county_20_50 = buildings_county[
        (buildings_county["year_built"] >= initial_year)
        & (buildings_county["building_type"].isin(residential_types))
    ]

    # Log the number of parcels in before removing duplicates
    initial_parcel_count = len(res_buildings_county_20_50)
    logging.info(
        f"Initial number of parcels in {fbp_id} new buildings summary: {initial_parcel_count}"
    )

    # Remove duplicate parcels since more the one building can be developled on a given parcel
    res_parcel_county_20_50 = res_buildings_county_20_50.drop_duplicates(
        subset="parcel_id", keep="first"
    )

    # Log the number of duplicates removed and the resulting number of parcels
    final_parcel_count = len(res_parcel_county_20_50)
    duplicates_removed = initial_parcel_count - final_parcel_count
    logging.info(
        f"Number of duplicates removed from {fbp_id} new buildings summary: {duplicates_removed}"
    )
    logging.info(
        f"Number of parcels remaining after removing duplicates: {final_parcel_count}"
    )

    # Define the parcel acre bins for assumed park generation from developed parcels
    bin_data = {
        "parcel_acre_bin": [
            "Less than 1",
            "1-2",
            "2-3",
            "3-5",
            "5-10",
            "10-20",
            "20-50",
            "50-100",
            "100-150",
            "150-200",
            "200-300",
            "300+",
        ],
        "share_devel_for_parks": [
            0,
            0,
            0.02,
            0.03,
            0.04,
            0.05,
            0.075,
            0.10,
            0.125,
            0.15,
            0.20,
            0.20,
        ],
        # "assumed_acres_of_bin": [
        #     0, 1.5, 2.5, 4, 7.5, 15, 35, 60, 125, 175, 250, 400 # Not currently used
        # ]
    }
    logging.info(
        f"Defining parcel acre bins for assumed park generation from developed parcels: {bin_data}"
    )
    bin_df = pd.DataFrame(bin_data, index=bin_data["parcel_acre_bin"])

    # Define bins and labels
    parcel_acre_bins = [0, 1, 2, 3, 5, 10, 20, 50, 100, 150, 200, 300, float("inf")]
    parcel_acre_bin_labels = bin_data["parcel_acre_bin"]

    # Bin the parcel acres column into a new column
    res_parcel_county_20_50["parcel_acres_bin"] = pd.cut(
        res_parcel_county_20_50["parcel_acres"],
        bins=parcel_acre_bins,
        labels=parcel_acre_bin_labels,
        right=False,
    )
    res_parcel_county_20_50 = res_parcel_county_20_50.merge(
        bin_df.drop(columns=["parcel_acre_bin"]),
        left_on="parcel_acres_bin",
        right_index=True,
        how="left",
    )

    # Calculate new park acres as share of parcel size
    logging.info(f"Calculating park acres generated from new residential development")
    res_parcel_county_20_50["new_park_acres_pt4"] = (
        res_parcel_county_20_50["share_devel_for_parks"]
        * res_parcel_county_20_50[
            "parcel_acres"
        ]  # Note: MG multiplies by assumed_acres_of_bin but since we have the actual parcel size, I think we should use that
    )

    # Keep only the necessary columns
    res_parcel_county_20_50 = res_parcel_county_20_50[
        [
            "parcel_id",
            "county_name",
            "parcel_acres",
            "parcel_acres_bin",
            "share_devel_for_parks",
            "new_park_acres_pt4",
        ]
    ]

    return res_parcel_county_20_50


def expand_urban_greening(
    BOX_DIR: pathlib.Path,
    rtp: str,
    modelrun_data: dict,
    output_path: str,
    append_output: bool,
):
    """
    Calculates urban greening metrics for PBA50+ off-model strategy EN6. The function generates a summary of urban park acreage,
    trail miles, and accessible open space acreage per 1000 residents for NP, DBP, and FBP.
        Args:
        - rtp (str): RTP2021 or RTP2025.
        - modelrun_data (dict): year -> {"parcel" -> parcel DataFrame}
        - output_path (str): File path for saving the output results
        - append_output (bool): True if appending output; False if writing
        Returns:
        - None: Saves the results to a CSV file.
    """
    logging.info("Calculating expand_urban_greening")
    if rtp == "RTP2021":
        logging.info("  RTP2021 is not supported - skipping")
        return

    # File paths
    EXISTING_PARK_ACRE_SUMMARY_FILE = (
        os.path.join(  # Generated from create_existing_park_summary.py
            BOX_DIR,
            "Modeling and Surveys",
            "Urban Modeling",
            "Spatial",
            "parks",
            "update_2025",
            "outputs",
            "existing_park_acres_summary.csv",
        )
    )

    NEW_PARK_ACRE_SUMMARY_FILE = os.path.join(
        BOX_DIR,
        "Modeling and Surveys",
        "Urban Modeling",
        "Spatial",
        "parks",
        "update_2025",
        "inputs",
        "MG_estimated_park_data",
        "estimated_new_park_acres_summary.csv",
    )

    TRAILS_OPEN_SPACE_SUMMARY_FILE = os.path.join(
        BOX_DIR,
        "Modeling and Surveys",
        "Urban Modeling",
        "Spatial",
        "parks",
        "update_2025",
        "inputs",
        "MG_estimated_park_data",
        "trails_open_space_summary2.csv",
    )

    # Load data
    logging.info("Loading existing and future park data")
    existing_park_acre_summary = pd.read_csv(EXISTING_PARK_ACRE_SUMMARY_FILE)
    parcel_devel_parks = generate_parks_from_new_development(
        fbp_id="PBA50Plus_Final_Blueprint_v65"
    )
    parcel_x_epc = modelrun_data[2050]["parcel"][
        ["parcel_id", "tract10_epc", "tract20_epc"]
    ]
    parcel_devel_parks = parcel_devel_parks.merge(
        parcel_x_epc, how="left", on="parcel_id"
    )

    # Load and join new park acres (Parts 1 through 3 of strategy EN6)
    new_park_acre_pt1_through_pt3 = pd.read_csv(NEW_PARK_ACRE_SUMMARY_FILE)

    # Summarize new park acres generated from development (Part 4 of strategy EN6)
    def summarize_new_park_acre_pt4(
        parcel_data, group_col, filter_col=None, filter_val=None
    ):
        if filter_col:
            parcel_data = parcel_data[parcel_data[filter_col] == filter_val]
        summary = (
            parcel_data.groupby(group_col)["new_park_acres_pt4"].sum().reset_index()
        )
        return summary

    new_park_acre_pt4_all_areas = summarize_new_park_acre_pt4(
        parcel_devel_parks, "county_name"
    )
    new_park_acre_pt4_all_areas["area_type"] = "All areas"

    new_park_acre_pt4_epc18 = summarize_new_park_acre_pt4(
        parcel_devel_parks, "county_name", "tract10_epc", 1
    )
    new_park_acre_pt4_epc18["area_type"] = "EPC_18"

    new_park_acre_pt4_epc22 = summarize_new_park_acre_pt4(
        parcel_devel_parks, "county_name", "tract20_epc", 1
    )
    new_park_acre_pt4_epc22["area_type"] = "EPC_22"

    new_park_acre_pt4_summary = pd.concat(
        [new_park_acre_pt4_all_areas, new_park_acre_pt4_epc18, new_park_acre_pt4_epc22],
        ignore_index=True,
    )

    # Merge Part 4 park acres with Parts 1 through 3
    new_park_acre_summary = new_park_acre_pt1_through_pt3.merge(
        new_park_acre_pt4_summary, on=["county_name", "area_type"], how="left"
    )

    # Calculate total new park acres: Parts 1-3 for DBP and Parts 1-4 for FBP
    def calculate_new_park_acres_from_plan(data, parts):
        data["new_park_acres"] = data[parts].sum(axis=1)
        data = data.drop(columns=parts)
        return data

    logging.info("Calculating new park acres for Draft Blueprint and Final Bluprint")
    parts = [
        "new_park_acres_pt1",
        "new_park_acres_pt2",
        "new_park_acres_pt3",
        "new_park_acres_pt4",
    ]
    new_park_acre_summary_dbp = calculate_new_park_acres_from_plan(
        new_park_acre_summary.copy(), parts[:-1]
    )
    new_park_acre_summary_fbp = calculate_new_park_acres_from_plan(
        new_park_acre_summary.copy(), parts
    )

    # Summarize population data for initial and horizon year, by region and by EPC
    logging.info("Loading and summarizing population data for initial and horizon year")
    pop_summary = pd.DataFrame()
    for year_idx, year in enumerate(sorted(modelrun_data.keys())):
        for area_type in ["", "EPC_22", "EPC_18"]:
            tazdata_df = modelrun_data[year]["TAZ1454"]
            if area_type == "EPC_22":
                epc_col = "tract20_epc"
                tazdata_df = tazdata_df.loc[tazdata_df[epc_col] == 1]
            elif area_type == "EPC_18":
                epc_col = "tract10_epc"
                tazdata_df = tazdata_df.loc[tazdata_df[epc_col] == 1]

            county_summary_df = (
                tazdata_df.groupby("COUNTY").agg({"TOTPOP": "sum"}).reset_index()
            )
            county_summary_df["year"] = year
            county_summary_df["area_type"] = (
                "All areas" if area_type == "" else area_type
            )
            logging.debug(
                f"  taz_county_summary_df for year {year}, segment {area_type}:\n{county_summary_df}"
            )
            pop_summary = pd.concat([pop_summary, county_summary_df])
    pop_summary = pop_summary.rename(
        columns={"COUNTY": "county_name", "TOTPOP": "tot_pop"}
    )

    # Subset pop summaries to 2023 and 2050
    pop_summary_2023 = pop_summary[pop_summary["year"] == 2023]
    pop_summary_2050 = pop_summary[pop_summary["year"] == 2050]

    # Create scenarios for NP 2023/2050, and DBP/FBP 2050
    logging.info("Creating scenarios for no project and plan")

    def create_scenario(
        existing_park_acres, new_park_acres, pop_data, alias, add_existing=False
    ):
        # Merge new park acres with population data
        scenario = new_park_acres.merge(
            pop_data, on=["county_name", "area_type"], how="left"
        )
        if add_existing:
            # Add existing park acres to new park acres (DBP and FBP)
            existing_cols = existing_park_acres[
                ["county_name", "area_type", "existing_park_acres"]
            ]
            scenario = scenario.merge(
                existing_cols,
                on=["county_name", "area_type"],
                how="left",
                suffixes=("", "_existing"),
            )
            scenario["tot_park_acres"] = scenario["new_park_acres"].fillna(
                0
            ) + scenario["existing_park_acres"].fillna(0)
        else:
            # Existing park acres is unchanged (NP)
            scenario = existing_park_acres.rename(
                columns={"existing_park_acres": "tot_park_acres"}
            )
            scenario = scenario.merge(
                pop_data, on=["county_name", "area_type"], how="left"
            )

        # Drop unnecessary columns
        scenario = scenario.drop(
            columns=["new_park_acres", "existing_park_acres"], errors="ignore"
        )
        scenario["alias"] = alias
        return scenario

    park_np23 = create_scenario(
        existing_park_acre_summary,
        existing_park_acre_summary,
        pop_summary_2023,
        "2023 No Project",
    )
    park_np50 = create_scenario(
        existing_park_acre_summary,
        existing_park_acre_summary,
        pop_summary_2050,
        "2050 No Project",
    )
    park_dbp50 = create_scenario(
        existing_park_acre_summary,
        new_park_acre_summary_dbp,
        pop_summary_2050,
        "2050 Draft Blueprint",
        add_existing=True,
    )
    park_fbp50 = create_scenario(
        existing_park_acre_summary,
        new_park_acre_summary_fbp,
        pop_summary_2050,
        "2050 Final Blueprint",
        add_existing=True,
    )

    # Prepare park acre summary for Tableau
    def prepare_tableau_summary(data, alias_col):
        summary = (
            data.groupby(["area_type"])
            .agg({"tot_park_acres": "sum", "tot_pop": "sum"})
            .reset_index()
        )
        summary["tot_pop_per_1k"] = summary["tot_pop"] / 1000
        summary["park_acre_per_1k"] = (
            summary["tot_park_acres"] / summary["tot_pop_per_1k"]
        )
        # summary = summary.rename(columns={'tot_park_acres': 'tot_urban_park_acres'})
        summary["alias"] = alias_col
        return summary

    region_park_np23 = prepare_tableau_summary(park_np23, "2023 No Project")
    region_park_np50 = prepare_tableau_summary(park_np50, "2050 No Project")
    region_park_dbp50 = prepare_tableau_summary(park_dbp50, "2050 Draft Blueprint")
    region_park_fbp50 = prepare_tableau_summary(park_fbp50, "2050 Final Blueprint")

    region_park_metrics = pd.concat(
        [region_park_np23, region_park_np50, region_park_dbp50, region_park_fbp50],
        ignore_index=True,
    )

    region_park_metrics = region_park_metrics.melt(
        id_vars=["area_type", "alias"],
        value_vars=["park_acre_per_1k"],
        var_name="metric_name",
        value_name="metric_value",
    )

    # Trails and open space (precalculated in a Workbook and read in)
    logging.info("Loading trails and open space data (pre-calculated in workbook)")
    trails_open_space = pd.read_csv(TRAILS_OPEN_SPACE_SUMMARY_FILE).merge(
        pop_summary, on=["county_name", "area_type", "year"], how="left"
    )
    region_trail_open_space_metrics = (
        trails_open_space.groupby(["year", "area_type", "modelrun_alias"])
        .agg({"open_space_acres": "sum", "trail_miles": "sum", "tot_pop": "sum"})
        .reset_index()
    )

    region_trail_open_space_metrics["tot_pop_per_1k"] = (
        region_trail_open_space_metrics["tot_pop"] / 1000
    )
    region_trail_open_space_metrics["open_space_acres_per_1k"] = (
        region_trail_open_space_metrics["open_space_acres"]
        / region_trail_open_space_metrics["tot_pop_per_1k"]
    )
    region_trail_open_space_metrics["trail_miles_per_1k"] = (
        region_trail_open_space_metrics["trail_miles"]
        / region_trail_open_space_metrics["tot_pop_per_1k"]
    )

    region_trail_open_space_metrics = region_trail_open_space_metrics.melt(
        id_vars=["year", "area_type", "modelrun_alias"],
        value_vars=["open_space_acres_per_1k", "trail_miles_per_1k"],
        var_name="metric_name",
        value_name="metric_value",
    )
    region_trail_open_space_metrics["alias"] = (
        region_trail_open_space_metrics["year"].astype(str)
        + " "
        + region_trail_open_space_metrics["modelrun_alias"]
    )
    region_trail_open_space_metrics = region_trail_open_space_metrics.drop(
        columns=["modelrun_alias", "year"]
    )

    # Combine and save
    final_park_metrics_summary = pd.concat(
        [region_park_metrics, region_trail_open_space_metrics], ignore_index=True
    )
    filename = "metrics_healthy1_urban_parks.csv"
    filepath = output_path / filename
    logging.info(f"Writing urban parks output formatted for tableau to {filepath}")
    final_park_metrics_summary.to_csv(filepath)


def urban_park_acres_in_workbook(
    BOX_DIR: pathlib.Path,
    rtp: str,
    modelrun_alias: str,
    modelrun_id: str,
    modelrun_data: dict,
    output_path: str,
    append_output: str,
):
    """
    This method works differently than the others since it works with an Excel workbook.
    1) If append_output == False, it copies the template workbook,
       PBA50+_DraftBlueprint_UrbanParksMetric.xlsx
       into the output_path as
       metrics_healthy1_UrbanParks.xlsx
    2) It fills in the relevant values (persons by county) into the workbook.

    The metrics_Healthy.twb directly reads from the local modified copy.

    Args:
        rtp (str): RTP2021 or RTP2025.
        modelrun_alias (str): Alias for the model run, used for labeling output.
        modelrun_id (str): Identifier for the model run.
        modelrun_data (dict): year -> {"parcel" -> parcel DataFrame, "county" -> county DataFrame }
        output_path (str or Path): The directory path to save the output XLSX file.
        append_output (bool): True if appending output; False if writing

    """
    logging.info("Calculating urban_park_acres")
    if rtp == "RTP2021":
        logging.info("  RTP2021 is not supported - skipping")
        return

    DEST_WORKBOOK = output_path / "metrics_healthy1_UrbanParks.xlsx"
    # first run -- copy template
    if append_output == False:
        SOURCE_WORKBOOK = (
            BOX_DIR
            / "Plan Bay Area 2050+/Performance and Equity/Plan Performance/Equity_Performance_Metrics/Healthy_Metrics_Templates/PBA50+_DraftBlueprint_UrbanParksMetric.xlsx"
        )
        shutil.copy(SOURCE_WORKBOOK, DEST_WORKBOOK)
        logging.info(f" Copied source workbook {SOURCE_WORKBOOK}")
        logging.info(f" to destination workbook {DEST_WORKBOOK}")

    workbook = openpyxl.load_workbook(filename=DEST_WORKBOOK)
    logging.debug("  Workbook sheetnames: {}".format(workbook.sheetnames))
    sheet = workbook["Dashboard"]
    logging.debug(f"  dimensions:{sheet.dimensions}")

    # summarize by year
    for year_idx, year in enumerate(sorted(modelrun_data.keys())):
        # and by Regionwide vs EPC
        for person_segment in ["", "EPC "]:

            # find: header columns for this year/person segment in worksheet
            # {year} {modelrun_alias} [EPC ]Population (Thousands)
            header_text = (
                f"{year} {modelrun_alias} {person_segment}Population (Thousands)"
            )
            logging.debug(f"  Looking for [{header_text}]")

            found_header_text = False
            row_num = None
            col_num = None
            for row_num in range(1, sheet.max_row + 1):
                for col_num in range(1, sheet.max_column + 1):
                    # logging.debug(f"{row_num},{col_num}: {sheet.cell(row_num, col_num).value}")
                    if sheet.cell(row_num, col_num).value == header_text:
                        logging.debug(f"  Found at {row_num},{col_num}")
                        found_header_text = True
                        break
                # break from row loop
                if found_header_text:
                    break

            if not found_header_text:
                if modelrun_alias in ["No Project", "DBP"]:
                    # this is a problem -- fail hard
                    logging.fatal(f"urban_park_acres not updated for {modelrun_alias}")
                    raise Exception(
                        f"urban_park_acres not updated for {modelrun_alias}"
                    )
                    raise ("")
                else:
                    logging.info(f"  => Didn't find {header_text} -- skipping")
                    continue

            # summarize the data for this
            tazdata_df = modelrun_data[year]["TAZ1454"]
            if person_segment == "EPC ":
                epc_col = "tract20_epc" if rtp == "RTP2025" else "tract10_epc"
                tazdata_df = tazdata_df.loc[tazdata_df[epc_col] == 1]

            taz_county_summary_df = tazdata_df.groupby("COUNTY").agg({"TOTPOP": "sum"})
            # convert to 1000s of population
            taz_county_summary_df["TOTPOP"] = taz_county_summary_df.TOTPOP / 1000
            logging.debug(f"  taz_county_summary_df:\n{taz_county_summary_df}")
            # fill in the workbook
            county_col_offset = -2
            if year_idx == 1:
                county_col_offset = -7

            for county_num in range(len(taz_county_summary_df)):
                # get county name from col-2
                county_name = sheet.cell(
                    row_num + county_num + 1, col_num + county_col_offset
                ).value
                logging.debug(
                    f"{county_name=} {row_num+county_num+1=} {col_num+county_col_offset=}"
                )
                # set the population
                sheet.cell(row_num + county_num + 1, col_num).value = (
                    taz_county_summary_df.loc[county_name, "TOTPOP"]
                )
            # skip one row and fill in the modelrun_id
            sheet.cell(row_num + len(taz_county_summary_df) + 2, col_num).value = (
                modelrun_id
            )

    # save and close the workbook
    workbook.save(DEST_WORKBOOK)
    logging.info(f"Saved {DEST_WORKBOOK}")
    workbook.close()

    # openpyxl can't make excel recalculate results; use xlwings for that
    # see https://stackoverflow.com/questions/36116162/python-openpyxl-data-only-true-returning-none/72901927#72901927
    excel_app = xlwings.App(visible=False)
    excel_book = excel_app.books.open(DEST_WORKBOOK)
    excel_book.save()
    excel_book.close()
    excel_app.quit()
    logging.info(f"Refreshed {DEST_WORKBOOK} with xlwings")


def ugb_development_share(
    rtp: str,
    modelrun_alias: str,
    modelrun_id: str,
    modelrun_data: dict,
    run_directory_path: pathlib.Path,
    output_path: pathlib.Path,
    append_output: bool,
):
    """
    Calculate and export the share of development that falls on parcels within the urban growth boundary
    (or is outside it but suitably low-density as to be rural in character).

    Parameters:
    - rtp (str): RTP2021 or RTP2025.
    - modelrun_alias (str): Alias for the model run, used for labeling output.
    - modelrun_id (str): Identifier for the model run.
    - modelrun_data (dict): year -> {"parcel" -> parcel DataFrame, "county" -> county DataFrame }
    - run_directory_path (Path): The directory path for this model run.
    - output_path (Path): The directory path to save the output CSV file.
    - append_output (bool): True if appending output; False if writing.
    """
    logging.info("Calculating ugb_development_share")

    # Guard clause: this metric is implemented for RTP2025 / PBA50+ only
    if rtp != "RTP2025":
        logging.info("  RTP2021 is not supported - skipping")
        return

    # Define a potentially very impactful constant used to convert residential units to non-residential sqft and vice versa
    SQFT_PER_UNIT = 1750  # close to the weighted average size of developer-model units in a recent BAUS run

    # Read in and select new buildings post 2020
    modelrun_name = modelrun_id
    # Sometimes the modelrun_id is a whole file path
    # Handle both forms of slashes in this field
    if "\\" in modelrun_id:
        modelrun_name = modelrun_id.split("\\")[-1]
    if "/" in modelrun_id:
        modelrun_name = modelrun_id.split("/")[-1]
    NEW_BUILDINGS_PATH = (
        pathlib.Path(run_directory_path)
        / f"core_summaries/{modelrun_name}_new_buildings_summary.csv"
    )
    logging.info(f"  Reading new_buildings_summary from {NEW_BUILDINGS_PATH}...")
    new_buildings = pd.read_csv(
        NEW_BUILDINGS_PATH,
        usecols=["parcel_id", "year_built", "building_sqft", "residential_units"],
        dtype={"parcel_id": int},
    )
    new_buildings = new_buildings[new_buildings["year_built"] > 2020]
    logging.debug(f"  {len(new_buildings)} buildings built after 2020")

    # Some residential buildings (from the development pipeline) have no building_sqft);
    # convert residential units to sqft equivalent so we can summarize "all development"
    new_buildings.loc[new_buildings["building_sqft"] == 0, "building_sqft"] = (
        new_buildings.loc[new_buildings["building_sqft"] == 0, "residential_units"]
        * SQFT_PER_UNIT
    )

    # We are interested in development on any parcel:
    # 1. outside the UGBs AND
    # 2. greater than 1 DU-equivalent per acre in 2050
    parcel_df = modelrun_data[2050]["parcel"].copy()
    parcel_df["du_equiv_per_acre"] = (
        parcel_df["residential_units"]
        + (parcel_df["non_residential_sqft"] / SQFT_PER_UNIT)
    ) / parcel_df["ACRES"]
    dense_greenfield_parcels = parcel_df.loc[
        (parcel_df["du_equiv_per_acre"] > 1.0) & (parcel_df["ugb_id"] != "UGB"),
        "parcel_id",
    ]

    # Calculate share of "all development" (in terms of building_sqft) that occurred on "dense greenfield parcels"
    total_development = new_buildings["building_sqft"].sum()
    dense_greenfield_development = new_buildings.loc[
        new_buildings["parcel_id"].isin(dense_greenfield_parcels), "building_sqft"
    ].sum()
    greenfield_development_pct = dense_greenfield_development / total_development

    # Add metadata, format, and export to CSV
    greenfield_development_df = pd.DataFrame(
        {
            "modelrun_id": modelrun_id,
            "modelrun_alias": modelrun_alias,
            "area_alias": "Regionwide",
            "area": "all",
            "development_in_urban_footprint_pct": 1 - greenfield_development_pct,
        },
        index=[0],
    )
    out_file = (
        pathlib.Path(output_path)
        / "metrics_healthy2_development_in_urban_footprint.csv"
    )
    greenfield_development_df.to_csv(
        out_file,
        mode="a" if append_output else "w",
        header=False if append_output else True,
        index=False,
    )
    logging.info(
        f"{'Appended' if append_output else 'Wrote'} {len(greenfield_development_df)} "
        + f"line{'s' if len(greenfield_development_df) > 1 else ''} to {out_file}"
    )


def slr_protection_v2(
    rtp, modelrun_alias, modelrun_id, modelrun_data, output_path, append_output
):
    """
    Calculates the percentage of households that are protected by sea level rise mitigation,
    as a percentage of all households at risk in sea level rise areas in the start year and 
    a percentage of all households in sea level rise areas that are EPCs.
    
    Parameters:
    - rtp (str): RTP2021 or RTP2025.
    - modelrun_alias (str): Alias for the model run, used for labeling output.
    - modelrun_id (str): Identifier for the model run.
    - modelrun_data (dict): year -> {"parcel" -> parcel DataFrame}
    - output_path (str): File path for saving the output results
    - append_output (bool): True if appending output; False if writing

    Writes metrics_healthy1_hazard_resilience_SLR.csv to output_path, appending if append_output is True. Columns are:
    - modelrun_alias
    - hazard
    - area_alias
    - area
    - protected_households_pct
    - protected_jobs_pct
    - protected_units_pct
    - hazard_alias
    """


    logging.info("Calculating slr_protection_v2")

    # Guard clause: this metric is implemented for RTP2025 / PBA50+ only
    if rtp != "RTP2025":
        logging.info("  RTP2021 is not supported - skipping")
        return

    # since inundation flag is time invariant, we don't need to run this for more than one year.

    this_modelrun_alias = metrics_utils.classify_runid_alias(modelrun_alias)

    # get parcel data for the horizon year
    horizon_year = sorted(modelrun_data.keys())[-1]
    initial_year = sorted(modelrun_data.keys())[0]
    
    # get the baseyear parcels - this has our starting households which we use for the at risk universe
    parcels_df = modelrun_data[initial_year]["parcel"].copy()

    # problem: with interpolation, we always get the NP baseyear - which means 
    # we get the NP specific inundation values, regardless of this_modelrun_alias. 
    # this is a problem for the FBP and DBP runs, which have different inundation values.
    # we therefore head to the horizon year for scenario specific inundation values
    # and bring those back to our base year risk universe
    
    # get the inundation values from the horizon year 
    parcels_df_horizon = modelrun_data[horizon_year]["parcel"].copy()
    
    # get just the inundation values
    future_inundation = parcels_df_horizon.set_index('parcel_id').inundation
    
    # assign back to the base year parcels
    # this is a bit of a hack, but we need to get the inundation values from the horizon year
    # and assign them to the base year parcels
    parcels_df["inundation"] = parcels_df.parcel_id.map(future_inundation)
            
    # classify risk, protection universes
    parcels_df["slr_at_risk"] = parcels_df["inundation"].notna()
    parcels_df["slr_protected"] = parcels_df.inundation.isin([100])

    # only keep at risk parcels
    
    logging.info(f"  {parcels_df.inundation.value_counts()}")
    logging.info(f"  subsetting parcels_df to only those at risk")
    logging.info(f"  {parcels_df.slr_at_risk.value_counts()}")    
    logging.info(f"  {parcels_df.slr_protected.value_counts()}")
    
    parcels_df = parcels_df.loc[parcels_df.slr_at_risk]
    # now we have only parcels that are at risk
    logging.info(f"len(parcels_df)={len(parcels_df):,}")
    SUMMARY_YEARS = sorted(modelrun_data.keys())[0]

    area_alias = {
        'EPC_18': 'Equity Priority Communities 2018',
        'EPC_22': 'Equity Priority Communities 2022',
        'Region': 'Region'
    }
    # We are focused on just the initial year - 2023
    # Get the SLR risk and protected households for each area   

    summary_list = []
    for area in ['EPC_18','EPC_22','Region']:
        filter_condition = metrics_utils.PARCEL_AREA_FILTERS[rtp][area]
        if callable(filter_condition):  # Check if the filter is a function
            df_area = parcels_df.loc[filter_condition(parcels_df)]
        elif filter_condition == None: # e.g. region
            df_area = parcels_df
        logging.debug("area={} df_area len={:,}".format(area, len(df_area)))

        # Calculate the share and append to results
        protected_hh_share = round(df_area.query('slr_protected').tothh.sum() / 
                           df_area.query('slr_at_risk').tothh.sum(),
                           ndigits=3)

        protected_emp_share = round(df_area.query('slr_protected').totemp.sum() / 
                           df_area.query('slr_at_risk').totemp.sum(),
                           ndigits=3)

        protected_unit_share = round(df_area.query('slr_protected').residential_units.sum() / 
                           df_area.query('slr_at_risk').residential_units.sum(),
                           ndigits=3)

        summary_list.append({
            'modelrun_id'             : modelrun_id,
            'modelrun_alias'          : f"{modelrun_alias}",
            'hazard'                  : "SLR",
            'hazard_alias'            : "Sea Level Rise",
            'area_alias'              : area_alias[area],
            'area'                    : area,
            'protected_households_pct': protected_hh_share,
            'protected_jobs_pct'      : protected_emp_share,
            'protected_units_pct'     : protected_unit_share,
        })

    # Create the results DataFrame
    hh_share_df = pd.DataFrame(summary_list)

    filename = "metrics_healthy1_hazard_resilience_SLR.csv"
    filepath = output_path / filename

    hh_share_df.to_csv(filepath, mode='a' if append_output else 'w', header=False if append_output else True, index=False)
    logging.info("{} {:,} lines to {}".format("Appended" if append_output else "Wrote", len(hh_share_df), filepath))